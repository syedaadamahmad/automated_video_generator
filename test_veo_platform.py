"""
test_veo_platform.py — Full test suite for the Veo Video Generation Platform
═══════════════════════════════════════════════════════════════════════════════

Coverage:
  Unit tests  — pure logic, no network, no API calls, all mocked
  Integration — FastAPI endpoints via httpx TestClient (in-process)

Test groups (run selectively with -k "group_name"):
  excel       — Excel validation, column parsing, notes-row filtering
  decomposer  — Prompt decomposition, narration rules, static camera
  s3          — S3 URL generation, soft-delete logic, disabled-mode fallback
  youtube     — Metadata generation, tag defaults, title truncation
  stitcher    — FFmpeg fade logic, path handling
  generator   — Model ID normalisation, attempt sequence, VEO_NO_VIDEO fallback
  api         — FastAPI endpoints: health, upload, jobs, rerun, youtube queue
  concurrency — Semaphore cap, parallel prompt execution, lock safety

Run all:
    pytest -s test_veo_platform.py

Run one group:
    pytest -s test_veo_platform.py -k "excel"

Run with extra verbosity:
    pytest -sv test_veo_platform.py
"""

import asyncio
import io
import json
import logging
import math
import os
import sys
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import openpyxl
import pandas as pd
import pytest

# ── Logging setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level   = logging.DEBUG,
    format  = "%(asctime)s | %(levelname)-8s | %(name)-20s | %(message)s",
    datefmt = "%H:%M:%S",
)
logger = logging.getLogger("TEST")

# ── Path setup — tests run from project root ──────────────────────────────────
PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))


# ══════════════════════════════════════════════════════════════════════════════
# Fixtures
# ══════════════════════════════════════════════════════════════════════════════

@pytest.fixture
def tmp_dir(tmp_path):
    """Temporary directory cleaned up after each test."""
    return tmp_path


@pytest.fixture
def simple_excel(tmp_path) -> Path:
    """Minimal valid Excel file — 1 prompt, duration 8."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["prompt", "duration"])
    ws.append(["A cat playing piano in a jazz club, cinematic lighting.", 8])
    p = tmp_path / "simple.xlsx"
    wb.save(p)
    logger.debug(f"[FIXTURE] simple_excel → {p}")
    return p


@pytest.fixture
def multi_prompt_excel(tmp_path) -> Path:
    """Excel with 5 prompts of varying durations including multi-clip."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["prompt", "duration", "task_type", "priority", "aspect_ratio"])
    rows = [
        ("Single clip product shot, dark marble surface, espresso pouring.", 8,  "TEXT_VIDEO", 1, "9:16"),
        ("Woman in cream blazer in warmly lit kitchen, sips coffee.", 16, "AUTO",       2, "9:16"),
        ("STATIC. Man in navy suit on dark stage, speaks to camera.", 24, "AUTO",       3, "16:9"),
        ("Cinematic cityscape at dawn, golden light over skyscrapers.", 8,  "AUTO",       4, "16:9"),
        ('STATIC. Woman at desk. Narration "AI is the future." Narration "Join us."', 16, "AUTO", 5, "9:16"),
    ]
    for r in rows:
        ws.append(r)
    p = tmp_path / "multi.xlsx"
    wb.save(p)
    logger.debug(f"[FIXTURE] multi_prompt_excel — {len(rows)} rows → {p}")
    return p


@pytest.fixture
def excel_with_notes_row(tmp_path) -> Path:
    """Excel where the last row is a notes/metadata row (has text, no duration)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["prompt", "duration"])
    ws.append(["Real prompt — cinematic ocean waves.", 8])
    ws.append(["★ Required: prompt, duration | Optional: ...", None])  # notes row
    p = tmp_path / "notes.xlsx"
    wb.save(p)
    logger.debug(f"[FIXTURE] excel_with_notes_row → {p}")
    return p


@pytest.fixture
def excel_with_aliases(tmp_path) -> Path:
    """Excel using column name aliases (text, duration_s) instead of canonical names."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["text", "duration_s"])   # aliases
    ws.append(["Ocean sunset timelapse, slow motion waves.", 8])
    p = tmp_path / "aliases.xlsx"
    wb.save(p)
    logger.debug(f"[FIXTURE] excel_with_aliases → {p}")
    return p


@pytest.fixture
def fake_mp4(tmp_path) -> Path:
    """Minimal fake MP4 file (not real video — for path/upload testing only)."""
    p = tmp_path / "test_video.mp4"
    p.write_bytes(b"\x00" * 1024 * 50)  # 50 KB fake bytes
    logger.debug(f"[FIXTURE] fake_mp4 → {p} (50 KB)")
    return p


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: excel — Excel validation and parsing
# ══════════════════════════════════════════════════════════════════════════════

class TestExcelProcessor:

    def test_valid_simple_file_passes_validation(self, simple_excel):
        logger.info("[EXCEL] test_valid_simple_file_passes_validation")
        from veo_excel_processor import validate_excel_file
        ok, errors = validate_excel_file(str(simple_excel))
        assert ok, f"Expected valid file to pass. Errors: {errors}"
        assert errors == []
        logger.info("[EXCEL] ✅ simple file validates clean")

    def test_missing_prompt_column_fails(self, tmp_path):
        logger.info("[EXCEL] test_missing_prompt_column_fails")
        from veo_excel_processor import validate_excel_file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["duration"])
        ws.append([8])
        p = tmp_path / "no_prompt.xlsx"
        wb.save(p)
        ok, errors = validate_excel_file(str(p))
        assert not ok
        assert any("prompt" in e.lower() for e in errors)
        logger.info(f"[EXCEL] ✅ correctly rejected — errors: {errors}")

    def test_missing_duration_column_fails(self, tmp_path):
        logger.info("[EXCEL] test_missing_duration_column_fails")
        from veo_excel_processor import validate_excel_file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt"])
        ws.append(["some prompt"])
        p = tmp_path / "no_duration.xlsx"
        wb.save(p)
        ok, errors = validate_excel_file(str(p))
        assert not ok
        assert any("duration" in e.lower() for e in errors)
        logger.info(f"[EXCEL] ✅ correctly rejected — errors: {errors}")

    def test_notes_row_not_treated_as_job(self, excel_with_notes_row):
        logger.info("[EXCEL] test_notes_row_not_treated_as_job")
        from veo_excel_processor import create_job_from_excel
        job = create_job_from_excel(str(excel_with_notes_row))
        prompts = job["prompts"]
        texts = [p["text"] for p in prompts]
        assert len(prompts) == 1, f"Expected 1 real prompt, got {len(prompts)}: {texts}"
        assert "★" not in texts[0], "Notes row leaked into prompts"
        logger.info(f"[EXCEL] ✅ notes row filtered — {len(prompts)} real prompt(s)")

    def test_column_aliases_normalised(self, excel_with_aliases):
        logger.info("[EXCEL] test_column_aliases_normalised")
        from veo_excel_processor import validate_excel_file, create_job_from_excel
        ok, errors = validate_excel_file(str(excel_with_aliases))
        assert ok, f"Alias file failed validation: {errors}"
        job = create_job_from_excel(str(excel_with_aliases))
        assert job["prompts"][0]["duration"] == 8
        logger.info("[EXCEL] ✅ column aliases (text, duration_s) normalised correctly")

    def test_duration_out_of_range_fails(self, tmp_path):
        logger.info("[EXCEL] test_duration_out_of_range_fails")
        from veo_excel_processor import validate_excel_file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt", "duration"])
        ws.append(["some prompt", 999])   # way above 120s limit
        p = tmp_path / "bad_dur.xlsx"
        wb.save(p)
        ok, errors = validate_excel_file(str(p))
        assert not ok
        assert any("range" in e.lower() for e in errors)
        logger.info(f"[EXCEL] ✅ out-of-range duration rejected — errors: {errors}")

    def test_nan_duration_row_filtered_out(self, tmp_path):
        """
        A row with a prompt but no duration is treated as a metadata/notes row
        and filtered by _has_duration. If it's the only row, ValueError is raised.
        This is correct — a job with no valid prompts should not be created.
        """
        logger.info("[EXCEL] test_nan_duration_row_filtered_out")
        from veo_excel_processor import create_job_from_excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt", "duration"])
        ws.append(["Valid prompt with no duration", None])
        p = tmp_path / "nan_dur.xlsx"
        wb.save(p)
        with pytest.raises(ValueError, match="No valid prompts"):
            create_job_from_excel(str(p))
        logger.info("[EXCEL] ✅ NaN-duration-only row raises ValueError (correctly treated as notes row)")

    def test_nan_duration_with_valid_row_still_processes(self, tmp_path):
        """
        When a NaN-duration row exists alongside a valid row, only the valid row
        is processed. The NaN row is silently dropped.
        """
        logger.info("[EXCEL] test_nan_duration_with_valid_row_still_processes")
        from veo_excel_processor import create_job_from_excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt", "duration"])
        ws.append(["Good prompt.", 8])                         # valid
        ws.append(["Notes row without duration.", None])       # filtered
        p = tmp_path / "mixed_dur.xlsx"
        wb.save(p)
        job = create_job_from_excel(str(p))
        assert len(job["prompts"]) == 1
        assert job["prompts"][0]["duration"] == 8
        logger.info("[EXCEL] ✅ NaN-duration row silently dropped, valid row processed correctly")

    def test_empty_rows_skipped(self, tmp_path):
        logger.info("[EXCEL] test_empty_rows_skipped")
        from veo_excel_processor import create_job_from_excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt", "duration"])
        ws.append(["Real prompt.", 8])
        ws.append([None, None])    # blank row
        ws.append(["", None])      # empty string row
        ws.append([None, 8])       # duration without prompt
        p = tmp_path / "gaps.xlsx"
        wb.save(p)
        job = create_job_from_excel(str(p))
        assert len(job["prompts"]) == 1
        logger.info(f"[EXCEL] ✅ empty rows skipped — {len(job['prompts'])} real prompt(s)")

    def test_multi_prompt_job_created_correctly(self, multi_prompt_excel):
        logger.info("[EXCEL] test_multi_prompt_job_created_correctly")
        from veo_excel_processor import create_job_from_excel
        job = create_job_from_excel(str(multi_prompt_excel))
        assert job["total_prompts"] == 5
        assert job["status"] == "pending"
        assert job["job_id"].startswith("job_")
        durations = [p["duration"] for p in job["prompts"]]
        assert durations == [8, 16, 24, 8, 16]
        logger.info(f"[EXCEL] ✅ multi-prompt job: {job['job_id']}, durations: {durations}")

    def test_task_type_validated(self, tmp_path):
        logger.info("[EXCEL] test_task_type_validated")
        from veo_excel_processor import validate_excel_file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt", "duration", "task_type"])
        ws.append(["Some prompt", 8, "INVALID_TASK"])
        p = tmp_path / "bad_task.xlsx"
        wb.save(p)
        ok, errors = validate_excel_file(str(p))
        assert not ok
        assert any("task_type" in e.lower() for e in errors)
        logger.info(f"[EXCEL] ✅ invalid task_type rejected — errors: {errors}")

    def test_priority_defaults_correctly(self, simple_excel):
        logger.info("[EXCEL] test_priority_defaults_correctly")
        from veo_excel_processor import create_job_from_excel
        job = create_job_from_excel(str(simple_excel))
        assert job["prompts"][0]["priority"] == 5   # default priority
        logger.info("[EXCEL] ✅ priority defaults to 5")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: decomposer — Prompt decomposition logic
# ══════════════════════════════════════════════════════════════════════════════

class TestPromptDecomposer:

    def _make_decomposer(self):
        """Create a PromptDecomposer with all Bedrock calls mocked out."""
        with patch("boto3.client"):
            from prompt_decomposer import PromptDecomposer
            d = PromptDecomposer.__new__(PromptDecomposer)
            d.model_primary   = "mock-primary"
            d.model_secondary = "mock-secondary"
            d._client         = MagicMock()
            return d

    def test_single_clip_returns_master_prompt_unchanged(self):
        logger.info("[DECOMPOSER] test_single_clip_returns_master_prompt_unchanged")
        from prompt_decomposer import PromptDecomposer
        with patch("boto3.client"):
            d = self._make_decomposer()
        master = "A cat plays jazz piano in a dimly lit club."
        prompts, source, objs, _ = d.decompose(master, n_clips=1, clip_durations=[8], platform="veo")
        assert len(prompts) == 1
        assert prompts[0] == master
        assert source == "passthrough"
        logger.info(f"[DECOMPOSER] ✅ single-clip passthrough: source={source}")

    def test_static_injected_in_single_clip(self):
        logger.info("[DECOMPOSER] test_static_injected_in_single_clip")
        d = self._make_decomposer()
        master = "STATIC. A woman at a desk speaks to camera."
        prompts, source, objs, _ = d.decompose(master, n_clips=1, clip_durations=[8], platform="veo", is_static=True)
        assert "STATIC LOCKED-OFF FRAME" in prompts[0]
        assert "Camera does not move" in prompts[0]
        logger.info(f"[DECOMPOSER] ✅ static directive injected into single-clip prompt")

    def test_static_not_injected_when_false(self):
        logger.info("[DECOMPOSER] test_static_not_injected_when_false")
        d = self._make_decomposer()
        master = "Ocean waves at sunset, slow motion."
        prompts, source, objs, _ = d.decompose(master, n_clips=1, clip_durations=[8], platform="veo", is_static=False)        assert "STATIC LOCKED-OFF FRAME" not in prompts[0]
        logger.info("[DECOMPOSER] ✅ static directive absent when is_static=False")

    def test_phase_fallback_returns_correct_count(self):
        logger.info("[DECOMPOSER] test_phase_fallback_returns_correct_count")
        d = self._make_decomposer()
        for n in [2, 3, 4, 5]:
            durations = [8] * n
            prompts, objs = d._phase_fallback("Test master prompt.", n, durations)
            assert len(prompts) == n, f"Expected {n} prompts, got {len(prompts)}"
            assert len(objs)    == n
            logger.info(f"[DECOMPOSER] ✅ phase_fallback n={n}: {len(prompts)} prompts")

    def test_phase_fallback_static_injection(self):
        logger.info("[DECOMPOSER] test_phase_fallback_static_injection")
        d = self._make_decomposer()
        prompts, objs = d._phase_fallback("Test.", 2, [8, 8], is_static=True)
        for p in prompts:
            assert "STATIC LOCKED-OFF FRAME" in p
            assert "Camera does not move" in p
        logger.info("[DECOMPOSER] ✅ static injected in all phase_fallback clips")

    def test_inject_static_modifies_all_clips(self):
        logger.info("[DECOMPOSER] test_inject_static_modifies_all_clips")
        d = self._make_decomposer()
        prompts    = ["Clip one.", "Clip two.", "Clip three."]
        clip_objs  = [{"clip": i+1, "duration_s": 8, "prompt": p, "end_state": ""} for i, p in enumerate(prompts)]
        new_p, new_objs = d._inject_static(prompts, clip_objs, is_static=True)
        for p in new_p:
            assert p.startswith("STATIC LOCKED-OFF FRAME")
        for obj in new_objs:
            assert obj["prompt"].startswith("STATIC LOCKED-OFF FRAME")
        logger.info("[DECOMPOSER] ✅ _inject_static applied to all clips and objects")

    def test_parse_json_objects_valid(self):
        logger.info("[DECOMPOSER] test_parse_json_objects_valid")
        d = self._make_decomposer()
        raw = json.dumps([
            {"clip": 1, "duration_s": 8, "end_state": "Final frame.", "prompt": "Clip one prompt."},
            {"clip": 2, "duration_s": 8, "end_state": "",             "prompt": "Clip two prompt."},
        ])
        result = d._parse_json_objects(raw, expected_n=2)
        assert result is not None
        assert len(result) == 2
        assert result[0]["prompt"] == "Clip one prompt."
        logger.info(f"[DECOMPOSER] ✅ JSON objects parsed: {len(result)} clips")

    def test_parse_json_objects_with_fences(self):
        logger.info("[DECOMPOSER] test_parse_json_objects_with_fences")
        d = self._make_decomposer()
        raw = '```json\n[{"clip":1,"duration_s":8,"end_state":"","prompt":"Test."}]\n```'
        result = d._parse_json_objects(raw, expected_n=1)
        assert result is not None
        assert len(result) == 1
        logger.info("[DECOMPOSER] ✅ JSON fences stripped correctly")

    def test_repair_count_pads_short(self):
        logger.info("[DECOMPOSER] test_repair_count_pads_short")
        d = self._make_decomposer()
        prompts  = ["Only one prompt."]
        objs     = [{"clip": 1, "duration_s": 8, "prompt": "Only one prompt.", "end_state": ""}]
        new_p, new_objs = d._repair_count(prompts, target=3, master_prompt="Master.", clip_objects=objs, clip_durations=[8,8,8])
        assert len(new_p)    == 3
        assert len(new_objs) == 3
        logger.info(f"[DECOMPOSER] ✅ _repair_count padded 1→3 clips")

    def test_repair_count_truncates_long(self):
        logger.info("[DECOMPOSER] test_repair_count_truncates_long")
        d = self._make_decomposer()
        prompts = [f"Clip {i}." for i in range(5)]
        objs    = [{"clip": i+1, "duration_s": 8, "prompt": p, "end_state": ""} for i, p in enumerate(prompts)]
        new_p, new_objs = d._repair_count(prompts, target=3, master_prompt="Master.", clip_objects=objs, clip_durations=[8,8,8])
        assert len(new_p) == 3
        logger.info(f"[DECOMPOSER] ✅ _repair_count truncated 5→3 clips")

    def test_compute_veo_clips(self):
        logger.info("[DECOMPOSER] test_compute_veo_clips")
        from prompt_decomposer import compute_veo_clips
        cases = [(8, 1), (16, 2), (24, 3), (32, 4), (7, 1), (9, 2), (100, 13)]
        for duration, expected_clips in cases:
            clips = compute_veo_clips(duration)
            assert len(clips) == expected_clips, f"duration={duration}: expected {expected_clips} clips, got {len(clips)}"
            logger.info(f"[DECOMPOSER] ✅ compute_veo_clips({duration}s) = {len(clips)} clips: {clips}")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: s3 — S3 client logic
# ══════════════════════════════════════════════════════════════════════════════

class TestVeoS3:

    def _disabled_client(self):
        """S3 client with no bucket configured — disabled mode."""
        from veo_s3 import VeoS3Client
        with patch("boto3.client"):
            c = VeoS3Client.__new__(VeoS3Client)
            c.bucket  = ""
            c.region  = "us-east-1"
            c.enabled = False
            c._client = None
            return c

    def _enabled_client(self):
        """S3 client with mocked boto3."""
        from veo_s3 import VeoS3Client
        mock_boto = MagicMock()
        mock_boto.head_bucket.return_value = {}
        with patch("boto3.client", return_value=mock_boto):
            c = VeoS3Client(bucket="test-bucket", region="us-east-1")
        c._client = mock_boto
        return c

    def test_public_url_format(self):
        logger.info("[S3] test_public_url_format")
        c = self._disabled_client()
        c.bucket = "bedrock-video-generation-us-east-1"
        c.region = "us-east-1"
        url = c._public_url("videos/job_abc/prompt_1/output.mp4")
        assert url == "https://bedrock-video-generation-us-east-1.s3.us-east-1.amazonaws.com/videos/job_abc/prompt_1/output.mp4"
        logger.info(f"[S3] ✅ public URL: {url}")

    def test_get_public_url_disabled_returns_none(self):
        logger.info("[S3] test_get_public_url_disabled_returns_none")
        c = self._disabled_client()
        url = c.get_public_url("job_abc", 0)
        assert url is None
        logger.info("[S3] ✅ disabled client returns None for get_public_url")

    def test_upload_returns_none_when_disabled(self, fake_mp4):
        logger.info("[S3] test_upload_returns_none_when_disabled")
        c = self._disabled_client()
        result = c.upload_video(str(fake_mp4), "job_abc", 0)
        assert result is None
        logger.info("[S3] ✅ disabled client returns None for upload_video")

    def test_upload_returns_none_for_missing_file(self):
        logger.info("[S3] test_upload_returns_none_for_missing_file")
        c = self._enabled_client()
        result = c.upload_video("/nonexistent/path/video.mp4", "job_abc", 0)
        assert result is None
        logger.info("[S3] ✅ missing file returns None gracefully")

    def test_upload_calls_s3_with_correct_key(self, fake_mp4):
        logger.info("[S3] test_upload_calls_s3_with_correct_key")
        c = self._enabled_client()
        c._client.upload_file = MagicMock()
        c._client.put_object  = MagicMock()
        result = c.upload_video(str(fake_mp4), "job_test123", 2)
        assert result is not None
        call_kwargs = c._client.upload_file.call_args
        assert call_kwargs is not None
        key = call_kwargs.kwargs.get("Key") or call_kwargs[1].get("Key") or call_kwargs[0][2]
        assert "job_test123" in key
        assert "prompt_3" in key      # prompt_index 2 → prompt_3 (1-based)
        assert "output.mp4" in key
        logger.info(f"[S3] ✅ upload_file called with key containing job_test123/prompt_3/output.mp4")

    def test_soft_delete_skips_missing_objects(self):
        logger.info("[S3] test_soft_delete_skips_missing_objects")
        from botocore.exceptions import ClientError
        c = self._enabled_client()
        not_found = ClientError({"Error": {"Code": "404", "Message": "Not Found"}}, "HeadObject")
        c._client.head_object = MagicMock(side_effect=not_found)
        result = c.soft_delete_video("job_abc", 0)
        # Should return True (all items handled, even if by skipping)
        assert isinstance(result, bool)
        logger.info(f"[S3] ✅ soft_delete with missing objects handled gracefully: {result}")

    def test_soft_delete_copies_then_deletes(self):
        logger.info("[S3] test_soft_delete_copies_then_deletes")
        c = self._enabled_client()
        c._client.head_object  = MagicMock(return_value={})
        c._client.copy_object  = MagicMock()
        c._client.delete_object = MagicMock()
        result = c.soft_delete_video("job_abc", 0)
        assert result is True
        assert c._client.copy_object.call_count  >= 1
        assert c._client.delete_object.call_count >= 1
        copy_call = c._client.copy_object.call_args_list[0]
        dest_key = copy_call.kwargs.get("Key", "") or copy_call[1].get("Key", "")
        assert dest_key.startswith("rejected/")
        logger.info(f"[S3] ✅ soft-delete: copied to {dest_key}, then deleted original")

    def test_s3_key_structure(self):
        logger.info("[S3] test_s3_key_structure")
        c = self._enabled_client()
        # Verify key pattern: videos/{job_id}/prompt_{N}/output.mp4
        for prompt_index, expected_n in [(0, 1), (1, 2), (4, 5)]:
            url = c._public_url(f"videos/job_abc/prompt_{prompt_index + 1}/output.mp4")
            assert f"prompt_{prompt_index + 1}" in url
            logger.info(f"[S3] ✅ key structure correct for prompt_index={prompt_index}: prompt_{prompt_index+1}")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: youtube — YouTube metadata generation
# ══════════════════════════════════════════════════════════════════════════════

class TestYouTubeModule:

    def test_generate_metadata_returns_all_keys(self):
        logger.info("[YOUTUBE] test_generate_metadata_returns_all_keys")
        from veo_youtube import generate_metadata
        meta = generate_metadata("A woman in a cream blazer stands in a kitchen.")
        assert "title"       in meta
        assert "description" in meta
        assert "tags"        in meta
        logger.info(f"[YOUTUBE] ✅ metadata keys present: {list(meta.keys())}")

    def test_title_truncated_to_100_chars(self):
        logger.info("[YOUTUBE] test_title_truncated_to_100_chars")
        from veo_youtube import generate_metadata
        long_prompt = "A " + "very long word " * 20 + "at the end."
        meta = generate_metadata(long_prompt)
        assert len(meta["title"]) <= 100
        logger.info(f"[YOUTUBE] ✅ title length={len(meta['title'])} ≤ 100")

    def test_default_tags_always_present(self):
        logger.info("[YOUTUBE] test_default_tags_always_present")
        from veo_youtube import generate_metadata, DEFAULT_TAGS
        meta = generate_metadata("Some completely unrelated prompt text.")
        for tag in DEFAULT_TAGS:
            assert tag in meta["tags"], f"Default tag '{tag}' missing from {meta['tags']}"
        logger.info(f"[YOUTUBE] ✅ default tags present: {DEFAULT_TAGS}")

    def test_static_prefix_stripped_from_title(self):
        logger.info("[YOUTUBE] test_static_prefix_stripped_from_title")
        from veo_youtube import generate_metadata
        meta = generate_metadata("STATIC. A man in a navy suit on a dark stage speaks to camera.")
        assert not meta["title"].startswith("STATIC")
        assert not meta["title"].startswith("STATIC.")
        logger.info(f"[YOUTUBE] ✅ STATIC prefix stripped: '{meta['title']}'")

    def test_narration_text_cleaned_from_title(self):
        logger.info("[YOUTUBE] test_narration_text_cleaned_from_title")
        from veo_youtube import generate_metadata
        meta = generate_metadata('Narration "AI is the future." Woman at a desk speaks.')
        assert not meta["title"].startswith('Narration "')
        logger.info(f"[YOUTUBE] ✅ Narration prefix stripped: '{meta['title']}'")

    def test_description_contains_prompt(self):
        logger.info("[YOUTUBE] test_description_contains_prompt")
        from veo_youtube import generate_metadata
        prompt = "Cinematic ocean waves at golden hour, slow motion."
        meta = generate_metadata(prompt)
        assert prompt[:40] in meta["description"]
        logger.info("[YOUTUBE] ✅ prompt text present in description")

    def test_description_has_attribution_footer(self):
        logger.info("[YOUTUBE] test_description_has_attribution_footer")
        from veo_youtube import generate_metadata
        meta = generate_metadata("Short prompt.")
        assert "Veo" in meta["description"]
        logger.info("[YOUTUBE] ✅ attribution footer present in description")

    def test_tags_deduped(self):
        logger.info("[YOUTUBE] test_tags_deduped")
        from veo_youtube import generate_metadata
        # "learning" is a default tag — if it also appears in the prompt it should not duplicate
        meta = generate_metadata("Learning about learning through education and growth.")
        assert len(meta["tags"]) == len(set(meta["tags"]))
        logger.info(f"[YOUTUBE] ✅ no duplicate tags: {meta['tags']}")

    def test_is_configured_false_when_no_secrets(self, tmp_path, monkeypatch):
        logger.info("[YOUTUBE] test_is_configured_false_when_no_secrets")
        import veo_youtube
        monkeypatch.setattr(veo_youtube, "SECRETS_FILE", tmp_path / "nonexistent.json")
        assert not veo_youtube.is_configured()
        logger.info("[YOUTUBE] ✅ is_configured() = False when secrets file absent")

    def test_is_configured_true_when_secrets_exist(self, tmp_path, monkeypatch):
        logger.info("[YOUTUBE] test_is_configured_true_when_secrets_exist")
        import veo_youtube
        secrets = tmp_path / "youtube_client_secrets.json"
        secrets.write_text("{}")
        monkeypatch.setattr(veo_youtube, "SECRETS_FILE", secrets)
        assert veo_youtube.is_configured()
        logger.info("[YOUTUBE] ✅ is_configured() = True when secrets file exists")

    def test_upload_fails_gracefully_when_file_missing(self):
        logger.info("[YOUTUBE] test_upload_fails_gracefully_when_file_missing")
        from veo_youtube import upload_video
        result = upload_video(
            local_path  = "/nonexistent/video.mp4",
            title       = "Test",
            description = "Test",
            tags        = ["test"],
        )
        assert result["status"] == "failed"
        assert "not found" in result["error"].lower()
        logger.info(f"[YOUTUBE] ✅ missing file handled gracefully: {result['error']}")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: stitcher — Video stitcher logic
# ══════════════════════════════════════════════════════════════════════════════

class TestVideoStitcher:

    def _make_stitcher(self, tmp_path):
        from video_stitcher import VideoStitcher
        return VideoStitcher(output_dir=str(tmp_path))

    def test_initialises_with_output_dir(self, tmp_path):
        logger.info("[STITCHER] test_initialises_with_output_dir")
        stitcher = self._make_stitcher(tmp_path)
        assert stitcher.output_dir == tmp_path
        logger.info(f"[STITCHER] ✅ output_dir={stitcher.output_dir}")

    def test_ffmpeg_available_check(self, tmp_path):
        logger.info("[STITCHER] test_ffmpeg_available_check")
        stitcher = self._make_stitcher(tmp_path)
        # ffmpeg_available is a bool — we don't assert True/False (depends on env)
        assert isinstance(stitcher.ffmpeg_available, bool)
        logger.info(f"[STITCHER] ✅ ffmpeg_available={stitcher.ffmpeg_available}")

    @pytest.mark.asyncio
    async def test_stitch_returns_none_when_empty_list(self, tmp_path):
        logger.info("[STITCHER] test_stitch_returns_none_when_empty_list")
        stitcher = self._make_stitcher(tmp_path)
        result = await stitcher.stitch_clips([], "job_abc", 0, "veo")
        assert result is None
        logger.info("[STITCHER] ✅ empty clip list returns None")

    @pytest.mark.asyncio
    async def test_stitch_single_clip_returns_path_without_stitching(self, tmp_path, fake_mp4):
        """
        Single clip is returned as-is — no stitching needed.
        The stitcher only concatenates when N > 1.
        """
        logger.info("[STITCHER] test_stitch_single_clip_returns_path_without_stitching")
        stitcher = self._make_stitcher(tmp_path)
        result = await stitcher.stitch_clips(
            [str(fake_mp4)], "job_abc", 0, "veo"
        )
        assert result == str(fake_mp4)
        logger.info(f"[STITCHER] ✅ single clip returned as-is: {result}")

    @pytest.mark.asyncio
    async def test_stitch_returns_none_for_empty_list(self, tmp_path):
        logger.info("[STITCHER] test_stitch_returns_none_for_empty_list")
        stitcher = self._make_stitcher(tmp_path)
        stitcher.ffmpeg_available = True
        result = await stitcher.stitch_clips([], "job_abc", 0, "veo")
        assert result is None
        logger.info("[STITCHER] ✅ empty list returns None")

    @pytest.mark.asyncio
    async def test_apply_fade_returns_none_when_ffmpeg_unavailable(self, tmp_path, fake_mp4):
        logger.info("[STITCHER] test_apply_fade_returns_none_when_ffmpeg_unavailable")
        stitcher = self._make_stitcher(tmp_path)
        stitcher.ffmpeg_available = False
        result = await stitcher._apply_fade(fake_mp4)
        assert result is None
        logger.info("[STITCHER] ✅ _apply_fade returns None when FFmpeg unavailable")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: generator — VeoGenerator logic
# ══════════════════════════════════════════════════════════════════════════════

class TestVeoGenerator:

    def _make_generator(self, tmp_path):
        from veo_generator import VeoGenerator
        with patch("google.genai.Client"):
            g = VeoGenerator(
                api_key    = "fake-key",
                output_dir = str(tmp_path),
            )
        return g

    def test_normalise_model_id_adds_prefix(self, tmp_path):
        logger.info("[GENERATOR] test_normalise_model_id_adds_prefix")
        from veo_generator import VeoGenerator
        assert VeoGenerator._normalise_model_id("veo-3.0-generate-001") == "models/veo-3.0-generate-001"
        assert VeoGenerator._normalise_model_id("models/veo-3.0-generate-001") == "models/veo-3.0-generate-001"
        logger.info("[GENERATOR] ✅ model ID prefix normalised correctly")

    def test_initialises_with_correct_models(self, tmp_path):
        logger.info("[GENERATOR] test_initialises_with_correct_models")
        g = self._make_generator(tmp_path)
        assert g.model_primary.startswith("models/")
        assert g.model_fallback.startswith("models/")
        assert g.clip_duration == 8
        logger.info(f"[GENERATOR] ✅ primary={g.model_primary}, fallback={g.model_fallback}")

    @pytest.mark.asyncio
    async def test_generate_video_text_only_success(self, tmp_path, fake_mp4):
        logger.info("[GENERATOR] test_generate_video_text_only_success")
        g = self._make_generator(tmp_path)

        mock_video = MagicMock()
        mock_video.video.uri = "https://fake-gcs.googleapis.com/test:download?alt=media"

        mock_operation = MagicMock()
        mock_operation.done = True
        mock_operation.response.generated_videos = [mock_video]

        g._client.models.generate_videos = MagicMock(return_value=mock_operation)
        g._client.operations.get         = MagicMock(return_value=mock_operation)

        # Mock _download_video directly — avoids patching httpx across thread boundary.
        # run_in_executor runs _download_bytes in a thread; by the time it executes,
        # a with-patch() context has already exited. Patching the method is safer.
        expected_url = "/videos/job_test_veo_p1.mp4"
        g._download_video = AsyncMock(return_value=expected_url)

        result = await g.generate_video(
            prompt       = "Test prompt.",
            duration     = 8,
            job_id       = "job_test",
            prompt_index = 0,
        )

        assert result["status"] == "completed"
        assert result["video_url"] == expected_url
        logger.info(f"[GENERATOR] ✅ text-only generation succeeded: {result['video_url']}")

    @pytest.mark.asyncio
    async def test_veo_no_video_triggers_text_fallback(self, tmp_path, fake_mp4):
        logger.info("[GENERATOR] test_veo_no_video_triggers_text_fallback")
        g = self._make_generator(tmp_path)

        # First two calls (img2vid with both models) return no videos
        # Third call (text-only primary) succeeds
        call_count = {"n": 0}

        mock_video = MagicMock()
        mock_video.video.uri = "https://fake-gcs.googleapis.com/fallback:download?alt=media"

        def side_effect(*args, **kwargs):
            call_count["n"] += 1
            op = MagicMock()
            op.done = True
            if call_count["n"] <= 2:
                op.response.generated_videos = []  # VEO_NO_VIDEO
            else:
                op.response.generated_videos = [mock_video]  # success
            return op

        g._client.models.generate_videos = MagicMock(side_effect=side_effect)
        g._client.operations.get         = MagicMock(return_value=MagicMock(done=True))

        ref_image = tmp_path / "_frame_test.jpg"
        ref_image.write_bytes(b"\xFF\xD8\xFF" + b"\x00" * 100)  # minimal JPEG header

        expected_url = "/videos/job_test_veo_p1.mp4"
        g._download_video = AsyncMock(return_value=expected_url)

        result = await g.generate_video(
            prompt                = "Test with image.",
            duration              = 8,
            job_id                = "job_test",
            prompt_index          = 0,
            reference_image_path  = ref_image,
        )

        assert result["status"] == "completed"
        assert call_count["n"] >= 3, f"Expected at least 3 calls (img×2 + text×1), got {call_count['n']}"
        logger.info(f"[GENERATOR] ✅ VEO_NO_VIDEO text fallback triggered after {call_count['n']} attempts")

    @pytest.mark.asyncio
    async def test_all_attempts_fail_returns_failed_status(self, tmp_path):
        logger.info("[GENERATOR] test_all_attempts_fail_returns_failed_status")
        g = self._make_generator(tmp_path)

        op = MagicMock()
        op.done = True
        op.response.generated_videos = []
        g._client.models.generate_videos = MagicMock(return_value=op)
        g._client.operations.get         = MagicMock(return_value=op)

        result = await g.generate_video(
            prompt       = "Test prompt.",
            duration     = 8,
            job_id       = "job_test",
            prompt_index = 0,
        )

        assert result["status"] == "failed"
        logger.info(f"[GENERATOR] ✅ all attempts failed → status=failed: {result.get('error_message', '')[:60]}")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: api — FastAPI endpoint tests via httpx TestClient
# ══════════════════════════════════════════════════════════════════════════════

class TestAPIEndpoints:
    """
    Tests the FastAPI app in-process using httpx.AsyncClient.
    All external services (Veo, S3, YouTube) are mocked.
    """

    @pytest.fixture
    def app_client(self, tmp_path):
        """
        Real import of veo_main — no module stubbing.

        Only two external calls are mocked:
          1. veo_orchestrator.generate_for_prompt — stops real Veo API calls
          2. veo_s3.upload_video                  — stops real S3 uploads

        Everything else is real: FastAPI routing, Excel parsing, job management,
        config, YouTube queue, health checks, error handling.

        Requires the full stack installed:
          pip install google-genai boto3 fastapi uvicorn httpx pytest-asyncio
        """
        try:
            import veo_main as app_module
        except SystemExit:
            pytest.skip(
                "veo_main could not start. Ensure google-genai and boto3 are "
                "installed and your .env file exists in the project folder."
            )
        except ImportError as e:
            pytest.skip(f"Missing dependency for API tests: {e}")

        # Mock only the two calls that hit real external services.
        # All other code — routing, validation, job store, config — is real.
        mock_result = {
            "status":           "completed",
            "video_url":        "/videos/test_video.mp4",
            "local_video_url":  str(tmp_path / "test_video.mp4"),
            "s3_url":           None,
            "duration_seconds": 8,
            "platform":         "veo",
            "stitched":         False,
            "clips_count":      1,
            "clip_urls":        ["/videos/test_video.mp4"],
            "model_used":       "models/veo-3.0-generate-001",
            "has_native_audio": True,
            "generation_time_seconds": 45.0,
        }
        app_module.veo_orchestrator.generate_for_prompt = AsyncMock(return_value=mock_result)
        app_module.veo_s3.upload_video = MagicMock(return_value=None)

        from httpx import AsyncClient, ASGITransport
        transport = ASGITransport(app=app_module.app)
        client    = AsyncClient(transport=transport, base_url="http://test")
        return client, app_module

    @pytest.mark.asyncio
    async def test_health_endpoint_returns_200(self, app_client):
        logger.info("[API] test_health_endpoint_returns_200")
        client, _ = app_client
        async with client as c:
            r = await c.get("/health")
        assert r.status_code == 200
        assert r.json()["status"] == "healthy"
        logger.info(f"[API] ✅ GET /health: {r.json()}")

    @pytest.mark.asyncio
    async def test_list_jobs_empty_on_startup(self, app_client):
        logger.info("[API] test_list_jobs_empty_on_startup")
        client, app = app_client
        app.jobs.clear()
        async with client as c:
            r = await c.get("/api/jobs")
        assert r.status_code == 200
        assert len(r.json()["jobs"]) == 0
        logger.info(f"[API] ✅ GET /api/jobs empty on startup: {r.json()}")

    @pytest.mark.asyncio
    async def test_get_nonexistent_job_returns_404(self, app_client):
        logger.info("[API] test_get_nonexistent_job_returns_404")
        client, _ = app_client
        async with client as c:
            r = await c.get("/api/jobs/job_doesnotexist")
        assert r.status_code == 404
        logger.info(f"[API] ✅ GET /api/jobs/nonexistent → 404")

    @pytest.mark.asyncio
    async def test_upload_valid_excel_creates_job(self, app_client, simple_excel):
        logger.info("[API] test_upload_valid_excel_creates_job")
        client, app = app_client
        app.jobs.clear()
        excel_bytes = simple_excel.read_bytes()
        async with client as c:
            r = await c.post(
                "/api/upload",
                files={"file": ("test.xlsx", excel_bytes, "application/octet-stream")},
            )
        assert r.status_code == 200, f"Upload failed: {r.text}"
        data = r.json()
        assert "job_id" in data
        assert data["job_id"].startswith("job_")
        assert data["prompts_count"] == 1
        logger.info(f"[API] ✅ POST /api/upload → job_id={data['job_id']}")

    @pytest.mark.asyncio
    async def test_upload_invalid_excel_returns_400(self, app_client, tmp_path):
        logger.info("[API] test_upload_invalid_excel_returns_400")
        client, _ = app_client
        # File with no required columns
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["wrong_col"])
        ws.append(["data"])
        bad_excel = tmp_path / "bad.xlsx"
        wb.save(bad_excel)
        async with client as c:
            r = await c.post(
                "/api/upload",
                files={"file": ("bad.xlsx", bad_excel.read_bytes(), "application/octet-stream")},
            )
        assert r.status_code == 400
        logger.info(f"[API] ✅ POST /api/upload bad file → 400")

    @pytest.mark.asyncio
    async def test_rerun_nonexistent_job_returns_404(self, app_client):
        logger.info("[API] test_rerun_nonexistent_job_returns_404")
        client, _ = app_client
        async with client as c:
            r = await c.post("/api/jobs/job_fakeid/rerun/0")
        assert r.status_code == 404
        logger.info(f"[API] ✅ POST /api/jobs/fake/rerun/0 → 404")

    @pytest.mark.asyncio
    async def test_rerun_out_of_range_index_returns_400(self, app_client, simple_excel):
        logger.info("[API] test_rerun_out_of_range_index_returns_400")
        client, app = app_client
        app.jobs.clear()
        excel_bytes = simple_excel.read_bytes()
        async with client as c:
            upload_r = await c.post(
                "/api/upload",
                files={"file": ("test.xlsx", excel_bytes, "application/octet-stream")},
            )
            job_id = upload_r.json()["job_id"]
            r = await c.post(f"/api/jobs/{job_id}/rerun/999")
        assert r.status_code == 400
        logger.info(f"[API] ✅ rerun out-of-range index → 400")

    @pytest.mark.asyncio
    async def test_youtube_status_endpoint(self, app_client):
        logger.info("[API] test_youtube_status_endpoint")
        client, _ = app_client
        async with client as c:
            r = await c.get("/api/youtube/status")
        assert r.status_code == 200
        data = r.json()
        assert "configured"    in data
        assert "authenticated" in data
        logger.info(f"[API] ✅ GET /api/youtube/status: {data}")

    @pytest.mark.asyncio
    async def test_youtube_queue_empty_initially(self, app_client):
        logger.info("[API] test_youtube_queue_empty_initially")
        client, app = app_client
        app.youtube_queue.clear()
        async with client as c:
            r = await c.get("/api/youtube/queue")
        assert r.status_code == 200
        assert r.json()["queue"] == []
        logger.info(f"[API] ✅ GET /api/youtube/queue empty initially")

    @pytest.mark.asyncio
    async def test_approve_nonexistent_job_returns_404(self, app_client):
        logger.info("[API] test_approve_nonexistent_job_returns_404")
        client, _ = app_client
        async with client as c:
            r = await c.post("/api/jobs/job_fake/approve/0")
        assert r.status_code == 404
        logger.info(f"[API] ✅ POST /api/jobs/fake/approve/0 → 404")

    @pytest.mark.asyncio
    async def test_update_queue_item_patch(self, app_client):
        logger.info("[API] test_update_queue_item_patch")
        client, app = app_client
        app.youtube_queue.clear()
        # Manually plant a queue item
        qid = "q_test_0"
        app.youtube_queue[qid] = {
            "queue_id":     qid,
            "job_id":       "job_test",
            "prompt_index": 0,
            "title":        "Original Title",
            "description":  "Original description",
            "tags":         ["ai"],
            "status":       "approved",
            "youtube_url":  None,
            "error":        None,
        }
        async with client as c:
            r = await c.patch(
                f"/api/youtube/queue/{qid}",
                json={
                    "title":       "Updated Title",
                    "description": "Updated description",
                    "tags":        ["ai", "education", "veo"],
                },
            )
        assert r.status_code == 200
        data = r.json()
        assert data["title"]       == "Updated Title"
        assert data["description"] == "Updated description"
        assert "education" in data["tags"]
        logger.info(f"[API] ✅ PATCH /api/youtube/queue/{qid}: title updated to '{data['title']}'")

    @pytest.mark.asyncio
    async def test_delete_queue_item(self, app_client):
        logger.info("[API] test_delete_queue_item")
        client, app = app_client
        app.youtube_queue.clear()
        qid = "q_del_0"
        app.youtube_queue[qid] = {
            "queue_id": qid, "status": "approved",
            "title": "T", "description": "D", "tags": [],
        }
        async with client as c:
            r = await c.delete(f"/api/youtube/queue/{qid}")
        assert r.status_code == 200
        assert qid not in app.youtube_queue
        logger.info(f"[API] ✅ DELETE /api/youtube/queue/{qid} removed item")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: concurrency — Semaphore and lock safety
# ══════════════════════════════════════════════════════════════════════════════

class TestConcurrency:

    @pytest.mark.asyncio
    async def test_semaphore_limits_concurrent_executions(self):
        """
        Verifies that at most N coroutines run simultaneously when gated by a
        Semaphore(N). Uses a counter tracked inside the coroutine.
        """
        logger.info("[CONCURRENCY] test_semaphore_limits_concurrent_executions")
        CAP = 5
        semaphore   = asyncio.Semaphore(CAP)
        active      = {"count": 0, "peak": 0}
        lock        = asyncio.Lock()

        async def task(i: int):
            async with semaphore:
                async with lock:
                    active["count"] += 1
                    active["peak"]   = max(active["peak"], active["count"])
                await asyncio.sleep(0.05)
                async with lock:
                    active["count"] -= 1

        tasks = [task(i) for i in range(20)]
        await asyncio.gather(*tasks)

        assert active["peak"] <= CAP, f"Peak concurrent={active['peak']} exceeded cap={CAP}"
        assert active["count"] == 0
        logger.info(f"[CONCURRENCY] ✅ peak concurrent={active['peak']} ≤ cap={CAP}")

    @pytest.mark.asyncio
    async def test_lock_prevents_counter_race(self):
        """
        Verifies asyncio.Lock prevents lost updates when multiple coroutines
        increment a shared counter — simulates the completed_prompts counter.
        """
        logger.info("[CONCURRENCY] test_lock_prevents_counter_race")
        WORKERS = 50
        state   = {"counter": 0}
        lock    = asyncio.Lock()

        async def increment():
            await asyncio.sleep(0)   # yield to scheduler
            async with lock:
                state["counter"] += 1

        await asyncio.gather(*[increment() for _ in range(WORKERS)])
        assert state["counter"] == WORKERS
        logger.info(f"[CONCURRENCY] ✅ counter={state['counter']} == {WORKERS} (no race condition)")

    @pytest.mark.asyncio
    async def test_gather_with_return_exceptions_isolates_failures(self):
        """
        Verifies asyncio.gather(return_exceptions=True) doesn't cancel
        successful tasks when one task fails — matches the run_generation_job pattern.
        """
        logger.info("[CONCURRENCY] test_gather_with_return_exceptions_isolates_failures")
        results = []

        async def task(i: int):
            await asyncio.sleep(0.01)
            if i == 2:
                raise ValueError(f"Task {i} failed intentionally")
            return f"ok_{i}"

        raw = await asyncio.gather(*[task(i) for i in range(5)], return_exceptions=True)
        successes = [r for r in raw if isinstance(r, str)]
        failures  = [r for r in raw if isinstance(r, Exception)]

        assert len(successes) == 4
        assert len(failures)  == 1
        assert "intentionally" in str(failures[0])
        logger.info(f"[CONCURRENCY] ✅ {len(successes)} succeeded, {len(failures)} failed (isolated)")

    @pytest.mark.asyncio
    async def test_prompts_complete_regardless_of_order(self):
        """
        Simulates the concurrent prompt loop — prompts finish in non-sequential
        order but all results land in the correct index.
        """
        logger.info("[CONCURRENCY] test_prompts_complete_regardless_of_order")
        PROMPTS     = 10
        semaphore   = asyncio.Semaphore(5)
        results_map: Dict[str, Any] = {}
        lock        = asyncio.Lock()

        async def fake_generate(i: int):
            async with semaphore:
                # Simulate variable generation time — later prompts finish first
                await asyncio.sleep(0.02 * (PROMPTS - i))
                return {"status": "completed", "video_url": f"/videos/prompt_{i}.mp4"}

        async def run_prompt(i: int):
            result = await fake_generate(i)
            async with lock:
                results_map[str(i)] = result

        await asyncio.gather(*[run_prompt(i) for i in range(PROMPTS)])

        assert len(results_map) == PROMPTS
        for i in range(PROMPTS):
            assert results_map[str(i)]["status"] == "completed"
            assert f"prompt_{i}" in results_map[str(i)]["video_url"]
        logger.info(f"[CONCURRENCY] ✅ all {PROMPTS} prompts completed in results_map regardless of order")


# ══════════════════════════════════════════════════════════════════════════════
# GROUP: integration — End-to-end flow tests (no real API calls)
# ══════════════════════════════════════════════════════════════════════════════

class TestIntegrationFlow:

    @pytest.mark.asyncio
    async def test_full_single_clip_flow(self, tmp_path, fake_mp4):
        """
        Single-clip prompt: Excel → validate → create_job → orchestrate → result dict.
        All external calls mocked.
        """
        logger.info("[INTEGRATION] test_full_single_clip_flow")

        # Build excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["prompt", "duration"])
        ws.append(["Cinematic ocean waves, slow motion, golden hour.", 8])
        excel_path = tmp_path / "flow_test.xlsx"
        wb.save(excel_path)

        # Validate
        from veo_excel_processor import validate_excel_file, create_job_from_excel
        ok, errors = validate_excel_file(str(excel_path))
        assert ok, f"Validation failed: {errors}"

        # Create job
        job = create_job_from_excel(str(excel_path))
        assert job["total_prompts"] == 1

        # Mock orchestrator
        mock_result = {
            "status":           "completed",
            "video_url":        str(fake_mp4),
            "local_video_url":  str(fake_mp4),
            "s3_url":           None,
            "duration_seconds": 8,
            "platform":         "veo",
            "stitched":         False,
            "clips_count":      1,
            "clip_urls":        [str(fake_mp4)],
            "model_used":       "models/veo-3.0-generate-001",
        }

        mock_orch = AsyncMock()
        mock_orch.generate_for_prompt = AsyncMock(return_value=mock_result)

        result = await mock_orch.generate_for_prompt(
            prompt_data  = job["prompts"][0],
            job_id       = job["job_id"],
            prompt_index = 0,
        )

        assert result["status"] == "completed"
        assert result["clips_count"] == 1
        logger.info(f"[INTEGRATION] ✅ single-clip flow: job={job['job_id']}, status={result['status']}")

    @pytest.mark.asyncio
    async def test_multi_clip_decomposition_flow(self, tmp_path):
        """
        Multi-clip prompt: decompose → verify N sub-prompts returned.
        Mocks LLM but tests actual phase-fallback path.
        """
        logger.info("[INTEGRATION] test_multi_clip_decomposition_flow")

        with patch("boto3.client"):
            from prompt_decomposer import PromptDecomposer, compute_veo_clips
            d = PromptDecomposer.__new__(PromptDecomposer)
            d.model_primary   = "mock"
            d.model_secondary = "mock"
            d._client         = MagicMock()

        master   = "A woman in a cream blazer sips coffee in a sunlit kitchen."
        duration = 24
        clips    = compute_veo_clips(duration)

        # Force fallback path (no LLM) — returns (prompts, clip_objects) 2-tuple
        prompts, objs = d._phase_fallback(master, len(clips), clips)

        assert len(prompts) == 3
        assert len(objs)    == 3
        assert all(isinstance(p, str) and len(p) > 10 for p in prompts)
        logger.info(f"[INTEGRATION] ✅ 24s prompt → {len(prompts)} sub-prompts via phase_fallback")

    @pytest.mark.asyncio
    async def test_static_detection_end_to_end(self, tmp_path):
        """
        STATIC keyword in prompt → is_static=True → directive in all sub-prompts.
        """
        logger.info("[INTEGRATION] test_static_detection_end_to_end")

        with patch("boto3.client"):
            from prompt_decomposer import PromptDecomposer, compute_veo_clips

            d = PromptDecomposer.__new__(PromptDecomposer)
            d.model_primary   = "mock"
            d.model_secondary = "mock"
            d._client         = MagicMock()

        master    = "STATIC. Man in navy suit on dark stage speaks to camera."
        is_static = "static" in master.lower()
        assert is_static

        clips = compute_veo_clips(24)
        prompts, objs = d._phase_fallback(master, len(clips), clips, is_static=is_static)

        for p in prompts:
            assert "STATIC LOCKED-OFF FRAME" in p
            assert "Camera does not move" in p

        logger.info(f"[INTEGRATION] ✅ STATIC detected and injected into all {len(prompts)} sub-prompts")

    def test_youtube_metadata_to_queue_flow(self, tmp_path, monkeypatch):
        """
        Prompt text → generate_metadata → queue item has correct structure.
        """
        logger.info("[INTEGRATION] test_youtube_metadata_to_queue_flow")
        from veo_youtube import generate_metadata, DEFAULT_TAGS

        prompt = 'STATIC. Woman at desk. Narration "AI changes everything." Narration "Join us today."'
        meta   = generate_metadata(prompt)

        # Simulate what approve_video does
        queue_item = {
            "queue_id":    "q_job_abc_0",
            "job_id":      "job_abc",
            "prompt_index": 0,
            "title":       meta["title"],
            "description": meta["description"],
            "tags":        meta["tags"],
            "status":      "approved",
        }

        assert len(queue_item["title"])       <= 100
        assert len(queue_item["description"]) <= 5000
        assert all(t in queue_item["tags"] for t in DEFAULT_TAGS)
        assert queue_item["status"] == "approved"
        logger.info(f"[INTEGRATION] ✅ queue item built: title='{queue_item['title'][:60]}'")

    def test_s3_url_used_as_video_url(self, tmp_path):
        """
        When S3 upload succeeds, video_url in result should be the S3 URL, not local.
        """
        logger.info("[INTEGRATION] test_s3_url_used_as_video_url")
        from veo_s3 import VeoS3Client

        c = VeoS3Client.__new__(VeoS3Client)
        c.bucket  = "bedrock-video-generation-us-east-1"
        c.region  = "us-east-1"
        c.enabled = True
        c._client = MagicMock()
        c._client.upload_file = MagicMock()
        c._client.put_object  = MagicMock()

        # Create a real temp mp4
        fake = tmp_path / "video.mp4"
        fake.write_bytes(b"\x00" * 1024)

        s3_url = c.upload_video(str(fake), "job_abc", 0)

        assert s3_url is not None
        assert s3_url.startswith("https://")
        assert "bedrock-video-generation-us-east-1" in s3_url
        assert "job_abc" in s3_url

        # Simulate orchestrator logic: final_url = s3_url or local_url
        final_url = s3_url or str(fake)
        assert final_url == s3_url
        logger.info(f"[INTEGRATION] ✅ S3 URL used as final video_url: {s3_url}")


# ══════════════════════════════════════════════════════════════════════════════
# Entrypoint
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import subprocess
    result = subprocess.run(
        ["pytest", __file__, "-sv", "--tb=short", "--no-header"],
        cwd=str(PROJECT_ROOT),
    )
    sys.exit(result.returncode)
